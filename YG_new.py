import sys
import serial
import time
import os
from PyQt5 import uic
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QLabel
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import *
from PyQt5.QtSerialPort import QSerialPort
from PyQt5.QtSerialPort import QSerialPortInfo
from PyQt5.QtCore import *
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QIODevice
from PyQt5.QtTest import QTest
from PyQt5.QtCore import QCoreApplication

import matplotlib
from matplotlib import pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

from abc import *

import openpyxl
from datetime import datetime
from openpyxl.chart import LineChart, Reference

import random
import YG_ui
from YG_ui import *
from matplotlibwidgetFile import matplotlibWidget
from matplotlibwidgetFile import MplCanvas

__platform__ = sys.platform


class myMplCanvas(MplCanvas):
    def __init__(self, parent, ma):
        super().__init__(parent, ma)
        self.pr_mouse_lo = [0, 0]
        self.mouse_location = [0, 0]
        self.xscale = 10000
        self.yscale = 1
        
    def adjScale(self):
        tmp = self.ax.axis()
        self.xscale = tmp[1] - tmp[0]
        self.yscale = tmp[3] - tmp[2]
    
    def mousePressEvent(self, ev):
        self.mouse_pressed = True
        if self.parent.objectName() == "sc1":
            self.ma.cl_or_wh_v_true()
        elif self.parent.objectName() == "sc2":
            self.ma.cl_or_wh_c_true()
        self.adjScale()
        
        self.pr_mouse_lo = [float((ev.x()-165)*(self.xscale/(986-165))), float((170-ev.y())*(self.yscale/170))]

    def mouseMoveEvent(self, ev):
        if self.mouse_pressed:
            self.adjScale()
            self.mouse_location = [float((ev.x()-165)*(self.xscale/(986-165))), float((170-ev.y())*(self.yscale/170))]
            dif = [self.pr_mouse_lo[0] - self.mouse_location[0], self.pr_mouse_lo[1] - self.mouse_location[1]]
            tmp = self.ax.axis()
            xmin = tmp[0] + dif[0]/10
            xmax = tmp[1] + dif[0]/10
            ymin = tmp[2] + dif[1]/10
            ymax = tmp[3] + dif[1]/10
            if xmin >= self.ma.mx:
                xmax = self.ma.mx
                xmin = xmax - self.xscale
            if xmin <= 0:
                xmin = 0
                xmax = self.xscale
            
            if self.parent.objectName() == "sc1":
                if ymax >= 85 + (85/3/2):
                    ymax = 85 + (85/3/2)
                    ymin = ymax - self.yscale
            elif self.parent.objectName() == "sc2":
                if ymax >= 165 + (165/3/2):
                    ymax = 165 + (165/3/2)
                    ymin = ymax - self.yscale
            self.ax.axis([xmin, xmax, ymin, ymax])
            if self.parent.objectName() == "sc1":
                self.ma.axi1 = [xmin, xmax, ymin, ymax]
                self.ma.draw_v_graph()
            elif self.parent.objectName() == "sc2":
                self.ma.axi2 = [xmin, xmax, ymin, ymax]
                self.ma.draw_c_graph()
            
    def mouseReleaseEvent(self, ev):
        self.mouse_pressed = False
        self.adjScale()
        dif = [self.mouse_location[0] - self.pr_mouse_lo[0], self.mouse_location[1] - self.pr_mouse_lo[1]]
        tmp = self.ax.axis()
        xmin = tmp[0] + dif[0]/10
        xmax = tmp[1] + dif[0]/10
        ymin = tmp[2] + dif[1]/10
        ymax = tmp[3] + dif[1]/10
        self.ax.axis([xmin, xmax, ymin, ymax])


    def wheelEvent(self, ev):
        x = ev.angleDelta()


class myMatplotlibWidget(matplotlibWidget):
    def __init__(self, parent, ma, name):
        super().__init__(parent)
        self.canvas = myMplCanvas(self, ma)
        self.parent = parent
        self.ma = ma
        self.setObjectName(name)
        
        self.vbl = QVBoxLayout()
        self.vbl.addWidget(self.canvas)
        self.setLayout(self.vbl)
        self.x_min = 0
        self.x_max = 0
        self.y_min = 0
        self.y_max = 0
        
    def draw(self):
        self.ma.draw_v_graph()
        self.ma.draw_c_graph()
        
    def scroll_bar(self):
        return [self.ma.main_ui.hsb_1.value(), self.ma.main_ui.hsb_2.value()]
    
    def cl_or_wh_v_true(self):
        self.ma.cl_or_wh_v_true()
        
    def cl_or_wh_c_true(self):
        self.ma.cl_or_wh_c_true()
        
    def cl_or_wh_v_false(self):
        self.ma.cl_or_wh_v_false()
        
    def cl_or_wh_c_false(self):
        self.ma.cl_or_wh_c_false()
    


class MyApp(QMainWindow):
    def __init__(self):
        super().__init__()
        QMainWindow.__init__(self)
        # UI 선언
        self.main_ui = Ui_Form()
        # UI 준비
        self.main_ui.setupUi(self)
        self.setupUi()
        
    def setupUi(self):
        
        self.setWindowTitle("EZDAQ")
        
        self.sc1 = myMatplotlibWidget(self.main_ui.vol_graph, self, "sc1")
        self.sc2 = myMatplotlibWidget(self.main_ui.cur_graph, self, "sc2")
        
        self.sc1.setGeometry(0, 0, 1131, 261)
        self.sc2.setGeometry(0, 0, 1131, 261)
        
        #쓰레드 생성
        self.re = Receive(self)
        self.ts = Test(self)
        
        #플래그
        #그래프1 클릭 or 휠
        self.cl_or_wh_v = False
        #그래프2 클릭 or 휠
        self.cl_or_wh_c = False
        
        self.ts_start = False
        
        #연결된 포트 번호 -1이라면 없음을 나타냄
        self.c = -1
        
        self.SerialPort = ' '
        
        self.arpt = 0
        
        self.range = 0
        
        #연결 후 경과된 시간
        self.tm = 0
        
        self.mx = 1
        self.mi = 0
        
        #그래프 배열
        self.g_gr = []
        self.vol_gr = []
        self.cur_gr = []
        #범위 저장
        self.cr_ax1 = []
        self.cr_ax2 = []
        
        #범위 계산용
        self.a_v = []
        self.a_c = []
        self.vol_arr = []
        self.cur_arr = []
        
        self.draw_v_graph()
        self.draw_c_graph()
        ##################
        tmp1 = self.sc1.canvas.ax.axis()
        xmin1 = tmp1[0]
        xmax1 = tmp1[1]
        ymin1 = tmp1[2]
        ymax1 = tmp1[3]
        
        tmp2 = self.sc2.canvas.ax.axis()
        xmin2 = tmp2[0]
        xmax2 = tmp2[1]
        ymin2 = tmp2[2]
        ymax2 = tmp2[3]
        
        self.axi1 = [xmin1, xmax1, ymin1, ymax1]
        self.axi2 = [xmin2, xmax2, ymin2, ymax2]
        ###################
        self.main_ui.hsb_1.setMinimum(0)
        self.main_ui.hsb_2.setMinimum(0)
        self.main_ui.vsb_1.setMaximum(85)
        self.main_ui.vsb_1.invertedControls = True
        self.main_ui.vsb_2.setMaximum(165)
        self.main_ui.vsb_2.invertedControls = True
        self.main_ui.hsb_1.setValue(self.main_ui.hsb_1.maximum())
        self.main_ui.hsb_2.setValue(self.main_ui.hsb_2.maximum())
        
        #포트 스캔 쓰레드 생성
        self.ck = Check(self)
        self.ck.start()
        
        #시리얼 설정
        self.serial = QSerialPort()
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        
        
        
        
        self.main_ui.hsb_1.actionTriggered.connect(self.hsb_1_action)
        self.main_ui.hsb_2.actionTriggered.connect(self.hsb_2_action)
        self.main_ui.vsb_1.actionTriggered.connect(self.vsb_1_action)
        self.main_ui.vsb_2.actionTriggered.connect(self.vsb_2_action)
        
        
        #연결 버튼 클릭
        self.main_ui.connect_c.clicked.connect(self.btn_clicked)
        #연결 해제 버튼 클릭
        self.main_ui.disconnect_c.clicked.connect(self.btn_clicked)
        #나가기 버튼 클릭
        self.main_ui.quit_c.clicked.connect(self.btn_clicked)
        #?REV: 버튼 클릭
        self.main_ui.Q_REV.clicked.connect(self.btn_clicked)
        #?SN: 버튼 클릭
        self.main_ui.Q_SN.clicked.connect(self.btn_clicked)
        #SET IRANGE 0
        self.main_ui.S_IRANGE_0.clicked.connect(self.btn_clicked)
        #SET IRANGE 1
        self.main_ui.S_IRANGE_1.clicked.connect(self.btn_clicked)
        #?IRANGE:
        self.main_ui.Q_IRANGE.clicked.connect(self.btn_clicked)
        #?MEAS:
        self.main_ui.Q_MEAS.clicked.connect(self.btn_clicked)
        #ARPT:0, 50, 1000, 5000, 10000
        self.main_ui.S_ARPT_0.clicked.connect(self.btn_clicked)
        self.main_ui.S_ARPT_50.clicked.connect(self.btn_clicked)
        self.main_ui.S_ARPT_1000.clicked.connect(self.btn_clicked)
        self.main_ui.S_ARPT_5000.clicked.connect(self.btn_clicked)
        self.main_ui.S_ARPT_10000.clicked.connect(self.btn_clicked)
        #AutoRange
        self.main_ui.autoRange_1.clicked.connect(self.btn_clicked)
        self.main_ui.autoRange_2.clicked.connect(self.btn_clicked)
        
        #커맨드 엔터 입력
        self.main_ui.command.returnPressed.connect(self.rPressed)
        
        self.main_ui.test.clicked.connect(self.testStart)
        
        #shutting down label 설정
        self.main_ui.lb.setAlignment(Qt.AlignCenter)
        self.main_ui.lb.setHidden(True)
        
        self.main_ui.saveLog.clicked.connect(self.saveExcel)
        self.main_ui.saveGraph.clicked.connect(self.saveExcelGraph)
    
    
    #테스트 함수
    def testStart(self):
        if not self.ts_start:
            self.ts.run()
            self.ts_start = True
        else:
            self.ts.stop()
            self.ts_start = False
    
    #연결
    def cnt(self, port_name):
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        if port_name:
            self.c = int(port_name[-1])
            self.serial.open(QIODevice.ReadWrite)
            #현재 IRANGE 확인
            self.serial.write("?IRANGE:\r".encode())
            
            self.s = self.serial.readLine()
            self.a= ''
            for i in self.s:
                self.a = self.a + i.decode()
            if self.a != '':
                self.range = int(self.a[8])
            
            #현재 ARPT 확인
            self.serial.write("?ARPT:\r".encode())
            
            return True
        else:
            return False
    
    #arpt 확인
    def ck_arpt(self):
        self.s = self.serial.readLine()
        self.a= ''
        self.b = ''
        for i in self.s:
            self.a = self.a + i.decode()
        if self.a != '':
            i = 6
            for d in self.a:
                if i > 0:
                    i -= 1
                    continue
                if ord(d) >= 48 and ord(d) <= 57:
                    self.b += d
                else:
                    break
        if self.b:
            self.arpt = int(self.b)
        
    #연결 확인
    def srl_IsOpen(self, port_name):
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        return self.serial.isOpen()
    
    #연결 해제
    def disconnect_serial(self):
        #정보 받아오는 쓰레드 중지
        self.re.stop()
        self.ts.stop()
        #연결 해제
        self.c = -1
        self.tm = 0
        return self.serial.close()
    
    #커멘드 입력
    def rPressed(self):
        self.s = self.command.text()
        self.st = self.s + '\r'
        self.serial.write(self.st.encode())
        self.s = (str)("<<<" + self.s)
        self.board_c.append(self.s)
        self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
        self.command.clear()
    
    #버튼 클릭
    def btn_clicked(self):
        self.btn_value = self.sender().text()
        #Connect
        if self.btn_value == 'Connect':
            #기존 연결 해제
            self.disconnect_serial()
            #연결하기
            self.con = self.cnt(self.main_ui.cbPort.currentText())
            if self.con:
                #정보 받아오는 쓰레드 생성
                self.re = Receive(self)
                self.re.start()
        
        #Disconnect
        elif self.btn_value == 'Disconnect':
            self.disconnect_serial()
            self.tm = 0
        
        #Quit
        elif self.btn_value == 'Quit':
            self.main_ui.lb.raise_()
            self.main_ui.lb.setHidden(False)
            self.main_ui.lb.repaint()
            self.disconnect_serial()
            #ck가 멈출 때까지 대기
            #포트 체크 중지
            self.ck.stop()
            QCoreApplication.instance().quit()
        
        #?REV:, ?SN:, ?IRANGE:, ?MEAS:
        elif self.btn_value == '?REV:' or self.btn_value == '?SN:' or self.btn_value == '?IRANGE:' or self.btn_value == '?MEAS:':
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        
        #SET IRANGE 0
        elif self.btn_value == 'IRANGE:0':
            self.range = 0
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #SET IRANGE 1
        elif self.btn_value == 'IRANGE:1':
            self.range = 1
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
        
        #ARPT:0
        elif self.btn_value == 'ARPT:0':
            self.arpt = 0
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #ARPT:50
        elif self.btn_value == 'ARPT:50':
            self.arpt = 50
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #ARPT:1000
        elif self.btn_value == 'ARPT:1000':
            self.arpt = 1000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #ARPT:5000
        elif self.btn_value == 'ARPT:5000':
            self.arpt = 5000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #ARPT:10000
        elif self.btn_value == 'ARPT:10000':
            self.arpt = 10000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.main_ui.board_c.append(self.s)
            self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
            self.main_ui.command.clear()
            
        #AutoRange
        elif self.btn_value == 'V_AutoRange':
            self.main_ui.autoRange_1.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
            self.cl_or_wh_v = False
        elif self.btn_value == 'C_AutoRange':
                self.main_ui.autoRange_2.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
                self.cl_or_wh_c = False
    
    #스크롤바
    def hsb_1_action(self):
        self.cl_or_wh_v = True
        self.main_ui.autoRange_1.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        tmp = self.sc1.canvas.ax.axis()
        xmin = tmp[0]
        xmax = tmp[1]
        ymin = tmp[2]
        ymax = tmp[3]
        xrange = xmax - xmin
        xmin = self.main_ui.hsb_1.value()
        xmax = xmin + xrange
        self.axi1 = [xmin, xmax, ymin, ymax]
        self.draw_v_graph()
        
    def vsb_1_action(self):
        self.cl_or_wh_c = True
        self.main_ui.autoRange_1.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        tmp = self.sc1.canvas.ax.axis()
        xmin = tmp[0]
        xmax = tmp[1]
        ymin = tmp[2]
        ymax = tmp[3]
        yrange = ymax - ymin
        ymin = self.main_ui.vsb_1.value()
        ymax = ymin + yrange
        self.axi1 = [xmin, xmax, ymin, ymax]
        self.draw_v_graph()
    
    def hsb_2_action(self):
        self.cl_or_wh_c = True
        self.main_ui.autoRange_2.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        tmp = self.sc2.canvas.ax.axis()
        xmin = tmp[0]
        xmax = tmp[1]
        ymin = tmp[2]
        ymax = tmp[3]
        xrange = xmax - xmin
        xmin = self.main_ui.hsb_2.value()
        xmax = xmin + xrange
        self.axi2 = [xmin, xmax, ymin, ymax]
        self.draw_c_graph()
        
    def vsb_2_action(self):
        self.cl_or_wh_c = True
        self.main_ui.autoRange_2.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        tmp = self.sc2.canvas.ax.axis()
        xmin = tmp[0]
        xmax = tmp[1]
        ymin = tmp[2]
        ymax = tmp[3]
        yrange = ymax - ymin
        ymin = self.main_ui.vsb_2.value()
        ymax = ymin + yrange
        self.axi2 = [xmin, xmax, ymin, ymax]
        self.draw_c_graph()
    
    #최근 가장 큰 값과 가장 작은 값 반환
    def vol_min_n_max(self, data):
        if self.arpt != 0:
            if len(self.vol_arr) > 10000/self.arpt:
                self.vol_arr.pop(0)
            self.vol_arr.append(data)
            a = min(self.vol_arr)
            b = max(self.vol_arr)
            c = ((b-a) * 1.25)/3/2
            if a == b:
                if a > 1.333:
                    return [a - 1, b + 1]
                else:
                    return[0, 1]
            else:
                return [a-c, b+c]
        else:
            return [0, 1]
        
    def cur_min_n_max(self, data):
        if self.arpt != 0:
            if len(self.cur_arr) > 10000/self.arpt:
                self.cur_arr.pop(0)
            self.cur_arr.append(data)
            a = min(self.cur_arr)
            b = max(self.cur_arr)
            c = ((b-a) * 1.25)/3/2
            if a == b:
                if a > 1.333:
                    return [a - 1, b+1]
                else:
                    return[0, 1]
            else:
                return [a-c, b+c]
        else:
            return[0, 1]
        
    #그래프 최대 최소 계산
    def graph_min_n_max(self):
        self.mx = self.tm + 2500
        if self.mx < 10000:
            self.mx = 10000
        self.mi = self.mx - 10000
        
    #그래프 오토레인지 범위 계산
    def v_auto_range(self):
        self.graph_min_n_max()
        if self.vol_gr:
            self.a_v = self.vol_min_n_max(self.vol_gr[-1])
            self.sc1.canvas.yscale = self.a_v[1] - self.a_v[0]
        else:
            self.a_v = [0, 1]
        self.sc1.canvas.ax.axis([self.mi, self.mx, self.a_v[0], self.a_v[1]])
    
    def c_auto_range(self):
        self.graph_min_n_max()
        if self.cur_gr:
            self.a_c = self.cur_min_n_max(self.cur_gr[-1])
            self.sc2.canvas.yscale = self.a_c[1] - self.a_c[0]
        else:
            self.a_c = [0, 1]
        self.sc2.canvas.ax.axis([self.mi, self.mx, self.a_c[0], self.a_c[1]])
    
    #그래프 그리기
    def draw_v_graph(self):
        self.sc1.canvas.ax.cla()
        if not self.cl_or_wh_v:
            self.v_auto_range()
        else:
            self.sc1.canvas.ax.axis(self.axi1)
        self.sc1.canvas.ax.plot(self.g_gr, self.vol_gr)
        self.sc1.canvas.draw()
        self.sc1.canvas.show()
        self.main_ui.hsb_1.setMaximum(self.tm)
        if not self.cl_or_wh_v:
            self.main_ui.hsb_1.setValue(self.main_ui.hsb_1.maximum())
    
    def draw_c_graph(self):
        self.sc2.canvas.ax.cla()
        if not self.cl_or_wh_c:
            self.c_auto_range()
        else:
            self.sc2.canvas.ax.axis(self.axi2)
        self.sc2.canvas.ax.plot(self.g_gr, self.cur_gr)
        self.sc2.canvas.draw()    
        self.sc2.canvas.show()
        self.main_ui.hsb_2.setMaximum(self.tm)
        if not self.cl_or_wh_c:
            self.main_ui.hsb_2.setValue(self.main_ui.hsb_2.maximum())
        
    def saveExcel(self):
        folder = QFileDialog.getExistingDirectory(self, 'select folder', '')
        ex = Excel(self, folder + '\\')
        
        ex.create()
        ex.write()
        ex.save()
    
    def saveExcelGraph(self):
        folder = QFileDialog.getExistingDirectory(self, 'select folder', '')
        ex = Excel(self, folder + '\\')
        
        ex.create()
        ex.write()
        ex.save_graph()
        
        #멤버 변수 접근
    def add_tm(self, tm):
        self.tm += tm
    
    def get_tm(self):
        return self.tm
    
    def get_c(self):
        return self.c
    
    def get_currentText(self):
        return self.main_ui.cbPort.currentText()
    
    def cb_remove(self, item):
        self.main_ui.cbPort.removeItem(item)
        
    def cb_findText(self, port_name):
        return self.main_ui.cbPort.findText(port_name)
    
    def cb_count(self):
        return self.main_ui.cbPort.count()
    
    def cb_addItem(self, item):
        self.main_ui.cbPort.addItem(item)
        
    def cb_itemText(self, text):
        return self.main_ui.cbPort.itemText(text)
    
    def cb_insertItem(self, num, name):
        self.main_ui.cbPort.insertItem(num, name)
    
    def get_serial_readLine(self):
        return self.serial.readLine()
    
    def get_board_c(self):
        return self.main_ui.board_c
    
    def get_btn_value(self):
        return self.btn_value
    
    def get_VOL(self):
        return self.main_ui.VOL
    
    def get_CUR(self):
        return self.main_ui.CUR
    
    def get_arpt(self):
        return self.arpt
    
    def get_g_gr(self):
        return self.g_gr
    
    def get_vol_gr(self):
        return self.vol_gr
    
    def get_cur_gr(self):
        return self.cur_gr
    
    def g_gr_append(self):
        self.g_gr.append(self.tm)
        
    def vol_gr_append(self, vol):
        self.vol_gr.append(vol)
        
    def cur_gr_append(self, cur):
        self.cur_gr.append(cur)
        
    def board_c_append(self, a):
        self.main_ui.board_c.append(a)
        
    def set_verScroll(self):
        self.main_ui.board_c.verticalScrollBar().setValue(self.main_ui.board_c.verticalScrollBar().maximum())
    
    def VOL_setText(self, text):
        self.main_ui.VOL.setText(text)
        
    def CUR_setText(self, text):
        self.main_ui.CUR.setText(text)
    
    def cl_or_wh_v_true(self):
        self.cl_or_wh_v = True
        self.main_ui.autoRange_1.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        
    def cl_or_wh_c_true(self):
        self.cl_or_wh_c = True
        self.main_ui.autoRange_2.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        
    def cl_or_wh_v_false(self):
        self.cl_or_wh_v = False
        self.main_ui.autoRange_1.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
        
    def cl_or_wh_c_false(self):
        self.cl_or_wh_c = False
        self.main_ui.autoRange_2.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
        
        
#들어오는 정보 감지 Thread
class Receive(QThread):
    def __init__(self, parent):
        QThread.__init__(self)
        self._status = False
        self.serial = serial
        self.re_w = True
        super().__init__(parent)
        self.parent = parent
        self.serial = QSerialPort()
        info = QSerialPortInfo(self.parent.get_currentText())
        self.serial.setPort(info)
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        
    #그래프 요소 추가
    def setGraph(self, vol, cur):
        self.parent.g_gr_append()
        self.parent.vol_gr_append(vol)
        self.parent.cur_gr_append(cur)
        
    def run(self):
        while self.re_w:
            self.s = self.parent.get_serial_readLine()
            if self.s:
                self.a= '>'
                for i in self.s:
                    self.a = self.a + i.decode('ascii')
                self.parent.board_c_append(self.a)
                self.parent.set_verScroll()
                if self.parent.get_btn_value() == '?MEAS:':
                    self.temp = self.a[5:]
                    self.tmp = self.temp.split(",")
                    
                    #vol display
                    while len(self.tmp[2]) < 6:
                        self.tmp[2] = self.tmp[2] + '0'
                    self.parent.VOL_setText(self.tmp[2][:3] + "," + self.tmp[2][3:])
                    
                    #cur display
                    while len(self.tmp[1]) < 6:
                        self.tmp[1] = self.tmp[1] + '0'
                    
                    self.parent.CUR_setText(self.tmp[1][:3] + "." + self.tmp[1][3:])
                    
                    
                if self.a[:6] == '>>ARPT' and self.parent.get_arpt() != 0:
                    self.temp = self.a[7:]
                    self.tmp = self.temp.split(",")
                    #vol display
                    self.tmp[3] = self.tmp[3].strip()
                    while len(self.tmp[3]) < 6:
                        self.tmp[3] = '0' + self.tmp[3]
                    self.parent.VOL_setText(self.tmp[3][:3] + "," + self.tmp[3][3:])
                    #cur display
                    self.tmp[2] = self.tmp[2].strip()
                    if int(self.tmp[2]) >= 0:
                        while len(self.tmp[2]) < 6:
                            self.tmp[2] = '0' + self.tmp[2]                  
                        self.parent.CUR_setText(self.tmp[2][:3] + "." + self.tmp[2][3:])
                    else:
                        while len(self.tmp[2]) < 6:
                            self.tmp[2] = self.tmp[2][0] + '0' + self.tmp[2][1:]                 
                        self.parent.CUR_setText(self.tmp[2][:3] + "." + self.tmp[2][3:])
                    
                    self.parent.add_tm(self.parent.get_arpt())
                    
                    self.setGraph(float(self.tmp[3]) / 1000, float(self.tmp[2]) / 1000)
                    #그래프 그리기
                    self.parent.draw_v_graph()
                    self.parent.draw_c_graph()
                    
            QTest.qWait(50)
            
    def stop(self):
        """
        QThread 종료
        """
        self.re_w = False
        
        self.wait()
        self.quit()

#포트 감지 Thread
class Check(QThread):
    def __init__(self, parent):
        QThread.__init__(self, parent)
        self._status = False
        self.serial = serial
        self.ck_w = True
        super().__init__(parent)
        self.parent = parent
    
    @staticmethod
    def get_port_path():
        """
        현재플래폼에 맞게 경로 또는 지정어를 반환
        :return:
        """
        return {"linux": '/dev/ttyS', "win32": 'COM'}[__platform__]
    
    # baudrate= 115200, data_bits=8, flow_control=None, parity=None, stop_bits=1
    def _open(self, port_name):
        """
        포트 연결 설정
        """
        self.serial = QSerialPort()
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        return self.serial.open(QIODevice.ReadWrite)
    
    def run(self):
        """
        255개의 포트를 열고 닫으면서 사용가능한 포트를 찾아서 입력
        """
        while self.ck_w:
            port_path = 'COM'
            for number in range(255):
                #ex) COM3
                port_name = port_path + str(number)
                #열리지 않고 이미 열려있지 않은 포트 삭제
                if (not self._open(port_name)) and number != self.parent.get_c():
                    self.parent.cb_remove(self.parent.cb_findText(port_name))
                    continue
                #포트가 열린다면 정렬 후 삽입
                if self._open(port_name):
                    #정렬
                    if self.parent.cb_count() != 0:
                        for num in range(0, self.parent.cb_count()):
                            if number > num:
                                continue
                            if number <= num:
                                self.parent.cb_insertItem(number, port_name)
                                break
                    elif self.parent.cb_count() == 0:
                        self.parent.cb_addItem(port_name)
                    self.serial.close()
            #sleep
            time.sleep(1)
        
    def stop(self):
        
        #QThread 종료
        
        self.ck_w = False
        self.wait()
        self.quit()
    
#작동 확인용 쓰레드
class Test(QThread):
    def __init__(self, parent):
        QThread.__init__(self)
        self._status = False
        self.serial = serial
        self.ck_w = True
        super().__init__(parent)
        self.parent = parent
        self.se = True
        
    def setGraph(self, vol, cur):
        self.parent.get_g_gr().append(self.parent.get_tm())
        self.parent.get_vol_gr().append(vol)
        self.parent.get_cur_gr().append(cur)
    
    def run(self):
        while self.se:
            self.parent.tm += 50
            self.setGraph(random.randrange(0, 86), random.randrange(0, 166))
            #그래프 그리기
            self.parent.draw_v_graph()
            self.parent.draw_c_graph()
            
            QTest.qWait(50)
            
    def stop(self):
        
        #QThread 종료
        
        self.se = False
        self.wait()
        self.quit()
    

class Excel():
    def __init__(self, parent, folder):
        self.parent = parent
        self.today = datetime.now()
        self.wb = None
        self.folder = folder
    
    def create(self):
        #생성
        self.wb = openpyxl.Workbook()
        #setting sheet name
        self.wb.active.title = "ezDAQ-Monitor"
        #setting sheet location
        self.w1 = self.wb["ezDAQ-Monitor"]
        #setting column name
        self.w1.cell(1, 1).value = 'ms'
        self.w1.cell(1, 2).value = 'v'
        self.w1.cell(1, 3).value = 'mA'
        
    def write(self):
        for t in range(0, len(self.parent.g_gr)):
            self.w1.cell(t+2, 1).value = self.parent.g_gr[t]
            self.w1.cell(t+2, 2).value = self.parent.vol_gr[t]
            self.w1.cell(t+2, 3).value = self.parent.cur_gr[t]
        
    #저장
    def save(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.new__filename = self.folder + today_s + ".xlsx"
        
        self.wb.save(self.new__filename)
    
    #그래프와 함께 저장
    def save_graph(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.new__filename = self.folder + "Graph" + today_s + ".xlsx"
        self.rd = self.wb
        sheet_list = self.rd.sheetnames
        s0 = self.rd[sheet_list[0]]
        
        chart1 = LineChart()
        chart2 = LineChart()
        #Data 범위 설정
        xval = Reference(s0, min_col=1 , max_col=1, min_row=2, max_row=len(self.parent.g_gr)+2)
        yval = Reference(s0, min_col = 2, max_col =2, min_row = 2, max_row = len(self.parent.g_gr)+2)
        chart1.add_data(yval, titles_from_data=True)
        chart1.set_categories(xval)
        chart1.style = 1
        
        chart1.height = 10
        chart1.width = (len(self.parent.g_gr)+2)
        s0.add_chart(chart1, "E5")
        
        yval = Reference(s0, min_col = 3, max_col =3, min_row = 2, max_row = len(self.parent.g_gr)+2)
        chart2.add_data(yval, titles_from_data=True)
        chart2.set_categories(xval)
        chart2.style = 1
        
        chart2.height = 10
        chart2.width = (len(self.parent.g_gr)+2)
        s0.add_chart(chart2, "E25")
        
        self.rd.save(self.new__filename)

    def read(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.filename = self.folder + today_s + ".xlsx"
        self.rd = openpyxl.load_workbook(self.filename, data_only=True)
        sheet_list = self.rd.sheetnames
        s0 = self.rd[sheet_list[0]]
        
        chart = LineChart()
        #Data 범위 설정
        xval = Reference(s0, min_col=1, max_col=1, min_row=2, max_row=len(self.parent.g)+2)
        yval = Reference(s0, min_col = 2, max_col =2, min_row = 2, max_row = len(self.parent.g)+2)
        chart.add_data(yval, titles_from_data=True)
        chart.set_categories(xval)
        chart.style = 1
        
        chart.height = 10
        chart.width = (len(self.parent.g_gr)+2)
        s0.add_chart(chart, "E5")
        
        self.new_filename = self.folder + today_s + ".xlsx"
        self.rd.save(self.new_filename)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    myApp = MyApp()
    myApp.show()
    
    app.exec_()