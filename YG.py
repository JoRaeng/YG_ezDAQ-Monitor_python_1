import sys
import serial
import time
import os
from PyQt5 import uic
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QLabel
from PyQt5.QtGui import *
from PyQt5.QtSerialPort import QSerialPort
from PyQt5.QtSerialPort import QSerialPortInfo
from PyQt5.QtCore import *
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QIODevice
from PyQt5.QtTest import QTest
from PyQt5.QtCore import QCoreApplication

import pyqtgraph as pg

from pyqtgraph import functions as fn

from abc import *

import openpyxl
from datetime import datetime
from openpyxl.chart import LineChart, Reference

#테스트용
import random

ui_path = r"YG_ui.ui"
'''
form_class = uic.loadUiType(ui_path)[0]
'''
__platform__ = sys.platform

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

form = resource_path("YG_ui.ui")
form_class = uic.loadUiType(form)[0]


class MyPlotWidget(pg.PlotWidget):

    def __init__(self, parent, name, hsb, vsb, arb, **kwargs):
        '''
        hsb 수평 스크롤 위젯
        vsb 수직 스크롤 위젯
        tm 가장 마지막 데이터가 들어온 시간(그래프 x값의 최댓값)
        cl_or_wh 그래프를 클릭하거나 드래그하면 True로 변함
        arb Auto Range 버튼 객체
        '''
        super().__init__(**kwargs)
        self.parent = parent
        self.name = name
        self.hsb = hsb
        self.vsb = vsb
        self.arb = arb
        self.vr = self.viewRange()
        self.verti = 0
        self.hori = 0
        self.le = 0
        self.ri = 0
        self.up = 0
        self.down = 0
        self.cen = [0.5, 25]
        self.cled = False
        
    def mousePressEvent(self, event):
        self.arb.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.cled = True
        #AutoRange 중지
        if self.name == '1':
            self.parent.cl_or_wh_v = True
        elif self.name == '2':
            self.parent.cl_or_wh_c = True
        super().mousePressEvent(event)
        #그래프 범위 풀기
        if self.name == '1':
            self.setLimits(xMin=0, xMax=self.parent.mx, yMin=0, yMax=85 + (85/3/2))
        elif self.name == '2':
            self.setLimits(xMin=0, xMax=self.parent.mx, yMin=0, yMax=165 + (165/3/2))
        
    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event)
        self.cled = False
        self.vr = self.viewRange()
    
    def mouseMoveEvent(self, event):
        if self.cled:
            self.vr = self.viewRange()
            super().mouseMoveEvent(event)
            #스크롤 위치 조정
            #가로
            halfh = (self.vr[0][1] - self.vr[0][0]) / 2

            self.hsb.setMinimum = int(halfh)
            self.hsb.setMaximum = int(self.parent.mx - halfh)
            self.hsb.setValue(int(self.vr[0][1] - halfh))
            self.hsb.setSliderPosition(int(self.vr[0][1] - halfh))
                
            #세로
            halfv = ((self.vr[1][1]) - int(self.vr[1][0])) / 2

            self.vsb.setMinimum = int(halfv)
            if self.name == '1':
                self.hsb.setMaximum = int(85 - halfv)
            elif self.name == '2':
                self.hsb.setMaximum = int(165 - halfv)
            self.vsb.setValue(int(self.vr[1][1] - halfv))
            self.vsb.setSliderPosition(int(self.vr[1][1] - halfv))
    
        
    def wheelEvent(self,event):
        #그래프 범위 풀기
        if self.name == '1':
            self.setLimits(xMin=0, xMax=self.parent.mx,minXRange=50, maxXRange=self.parent.mx, yMin=0, yMax=85 + (85/3/2), minYRange=0.001, maxYRange=85 + (85/3/2))
        elif self.name == '2':
            self.setLimits(xMin=0, xMax=self.parent.mx, minXRange=50, maxXRange=self.parent.mx, yMin=0, yMax=165 + (165/3/2), minYRange=0.001, maxYRange=165 + (165/3/2))
        
        super().wheelEvent(event)
        #AutoRange 중지
        if self.name == '1':
            self.parent.cl_or_wh_v = True
        elif self.name == '2':
            self.parent.cl_or_wh_c = True
        self.arb.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.vr = self.viewRange()
        
        #가로 스크롤 범위 설정
        #가로 화면에 나오는 x좌표 오른쪽 끝 - 왼쪽 끝 나누기 2
        
        half1 = (self.vr[0][1] - self.vr[0][0]) / 2
        
        #가로 스크롤 크기 조정
        self.hsb.setPageStep(int(half1))
        
        #가로 스크롤 범위 조정
        self.hsb.setMinimum(int(half1))
        if self.parent.mx - half1 <= half1:
            self.hsb.setMaximum(int(half1+1))
        else:
            self.hsb.setMaximum(int((self.parent.mx - half1)))
        
        #가로 스크롤 위치 설정
        self.hsb.setValue(int(self.vr[0][1] - half1))
        self.hsb.setSliderPosition(int(self.vr[0][1] - half1))
        
        #세로 화면에 나오는 y좌표 맨위 - 맨아래
        half2 = (self.vr[1][1] - self.vr[1][0]) / 2
        
        #세로 스크롤 크기 조정
        self.vsb.setPageStep(int(half2*8))
        
        #세로 스크롤 범위 조정
        self.vsb.setMinimum(0)
        
        #세로 스크롤 위치 설정
        self.vsb.setValue(int(self.vr[1][1] - half2))
        self.vsb.setSliderPosition(int(self.vr[1][1] - half2))
        
    
    def ret_vr(self):
        return self.viewRange()
    

class MyQLabel(QLabel):
    def __init__(self, parent, **kwargs):
        super().__init__(**kwargs)
        self.parent = parent
        self.setHidden(True)
    
    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event)
        if self.isHidden():
            self.setHidden(False)
        else:
            self.setHidden(True)

class MyApp(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        
        #mA문자 이미지 라벨에 부착
        pixmap = QPixmap('C:\ezDAQ-Monitor\mA.png')
        self.label.setPixmap(pixmap)
        
        #쓰레드 생성
        self.re = Receive(self)
        self.ts = Test(self)
        #플래그
        #그래프1 클릭 or 휠
        self.cl_or_wh_v = False
        #그래프2 클릭 or 휠
        self.cl_or_wh_c = False
        
        self.ts_start = False
        
        #연결된 포트 번호 -1이라면 없음을 나타냄
        self.c = -1
        
        self.SerialPort = ' '
        
        self.arpt = 0
        
        self.range = 0
        
        #연결 후 경과된 시간
        self.tm = 0
        
        self.mx = 1
        self.mi = 0
        
        #그래프 배열
        
        self.g = []
        self.vol_g_y = []
        self.cur_g_y = []
        
        self.vol_arr = []
        self.cur_arr = []
        
        self.a_v = []
        self.a_c = []
        
        #그래프 생성
        self.sc1 = MyPlotWidget(self, '1', self.hsb_1, self.vsb_1, self.autoRange_1)
        self.sc2 = MyPlotWidget(self, '2', self.hsb_2, self.vsb_2, self.autoRange_2)
        
        self.sc1.setParent(self.vol_graph)
        self.sc2.setParent(self.cur_graph)
        
        self.sc1.setGeometry(0, 0, 1131, 261)
        self.sc2.setGeometry(0, 0, 1131, 261)
        
        self.sc1.showGrid(x=True, y=True)
        self.sc2.showGrid(x=True, y=True)
        
        #스크롤바 설정
        self.vsb_1.setMinimum(0)
        self.vsb_1.setMaximum(85)
        self.vsb_1.setValue(0)
        self.vsb_1.setSliderPosition(0)
        
        self.vsb_2.setMinimum(0)
        self.vsb_2.setMaximum(165)
        self.vsb_2.setValue(0)
        self.vsb_2.setSliderPosition(0)
        
        self.hsb_1.setMinimum(0)
        self.hsb_1.setMaximum(self.mx)
        self.hsb_1.setValue(0)
        self.hsb_1.setSliderPosition(0)
        
        self.hsb_2.setMinimum(0)
        self.hsb_2.setMaximum(self.mx)
        self.hsb_2.setValue(0)
        self.hsb_2.setSliderPosition(0)
        
        self.hsb_1.setSingleStep(1)
        self.hsb_2.setSingleStep(1)
        
        self.hhalf1 = 100
        self.hhalf2 = 100
        self.vhalf1 = 43
        self.shalf2 = 83
        
        self.hsb_1.actionTriggered.connect(self.hsb_1_action)
        self.hsb_2.actionTriggered.connect(self.hsb_2_action)
        self.vsb_1.actionTriggered.connect(self.vsb_1_action)
        self.vsb_2.actionTriggered.connect(self.vsb_2_action)
        
        
        self.draw_v_gr()
        self.draw_c_gr()
        
        #포트 스캔 쓰레드 생성
        self.ck = Check(self)
        self.ck.start()
        
        #시리얼 설정
        self.serial = QSerialPort()
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        
        #연결 버튼 클릭
        self.connect_c.clicked.connect(self.btn_clicked)
        #연결 해제 버튼 클릭
        self.disconnect_c.clicked.connect(self.btn_clicked)
        #나가기 버튼 클릭
        self.quit_c.clicked.connect(self.btn_clicked)
        #?REV: 버튼 클릭
        self.Q_REV.clicked.connect(self.btn_clicked)
        #?SN: 버튼 클릭
        self.Q_SN.clicked.connect(self.btn_clicked)
        #SET IRANGE 0
        self.S_IRANGE_0.clicked.connect(self.btn_clicked)
        #SET IRANGE 1
        self.S_IRANGE_1.clicked.connect(self.btn_clicked)
        #?IRANGE:
        self.Q_IRANGE.clicked.connect(self.btn_clicked)
        #?MEAS:
        self.Q_MEAS.clicked.connect(self.btn_clicked)
        #ARPT:0, 50, 1000, 5000, 10000
        self.S_ARPT_0.clicked.connect(self.btn_clicked)
        self.S_ARPT_50.clicked.connect(self.btn_clicked)
        self.S_ARPT_1000.clicked.connect(self.btn_clicked)
        self.S_ARPT_5000.clicked.connect(self.btn_clicked)
        self.S_ARPT_10000.clicked.connect(self.btn_clicked)
        #AutoRange
        self.autoRange_1.clicked.connect(self.btn_clicked)
        self.autoRange_2.clicked.connect(self.btn_clicked)
        
        
        #커맨드 엔터 입력
        self.command.returnPressed.connect(self.rPressed)
        
        self.test.clicked.connect(self.testStart)
        
        #shutting down label 설정
        self.lb.setAlignment(Qt.AlignCenter)
        self.lb.setHidden(True)
        
        self.saveLog.clicked.connect(self.saveExcel)
        self.saveGraph.clicked.connect(self.saveExcelGraph)
        
    
    
    #테스트 함수
    def testStart(self):
        self.cl_or_wh_v = False
        self.cl_or_wh_c = False
        if not self.ts_start:
            self.ts.run()
            self.ts_start = True
        elif self.ts_start:
            self.ts.stop()
            self.ts_start = False
    
    #그래프 그리기
    def draw_v_gr(self):
        #clear() 로 이전에 그린 차트 제거함. 
        self.sc1.clear()
        self.set_limit()
        #그래프 그리기
        self.sc1.plot(self.g, self.vol_g_y)
        
        #스크롤바 재설정
        self.sc1.ret_vr()
        self.hsb_1.Maximum = int(self.mx)
        if not self.cl_or_wh_v:
            self.hsb_1.setValue(self.hsb_1.maximum())
            self.hsb_1.setPageStep(int(100000/self.mx))
        
    def draw_c_gr(self):
        #clear() 로 이전에 그린 차트 제거함. 
        self.sc2.clear()
        self.set_limit()
        #그래프 그리기
        self.sc2.plot(self.g, self.cur_g_y)
        
        #스크롤바 재설정
        self.hsb_2.Maximum = int(self.mx)
        if not self.cl_or_wh_c:
            self.hsb_2.setValue(self.hsb_2.maximum())
            self.hsb_2.setPageStep(int(100000/self.mx))
        
    #최근 200개중 가장 큰 값과 가장 작은 값 반환
    def vol_min_n_max(self, data):
        if self.arpt != 0:
            if len(self.vol_arr) > 10000/self.arpt:
                self.vol_arr.pop(0)
            self.vol_arr.append(data)
            a = min(self.vol_arr)
            b = max(self.vol_arr)
            c = ((b-a) * 1.25)/3/2
            if a == b:
                if a > 1.333:
                    return [a - 1, b + 1]
                else:
                    return[0, 1]
            else:
                return [a-c, b+c]
        else:
            return [0, 1]
    
    def cur_min_n_max(self, data):
        if self.arpt != 0:
            if len(self.cur_arr) > 10000/self.arpt:
                self.cur_arr.pop(0)
            self.cur_arr.append(data)
            a = min(self.cur_arr)
            b = max(self.cur_arr)
            c = ((b-a) * 1.25)/3/2
            if a == b:
                if a > 1.333:
                    return [a - 1, b+1]
                else:
                    return[0, 1]
            else:
                return [a-c, b+c]
        else:
            return[0, 1]
        
    #그래프 최대 최소 계산
    def graph_min_n_max(self):
        self.mx = self.tm + 2500
        if self.mx < 10000:
            self.mx = 10000
        self.mi = self.mx - 10000
    
    #그래프 범위
    def set_limit(self):
        self.graph_min_n_max()
        #이동
        if not self.cl_or_wh_v:
            if self.vol_g_y:
                self.a_v = self.vol_min_n_max(self.vol_g_y[-1])
            else:
                self.a_v = [0, 1]
            self.sc1.setLimits(xMin=self.mi, xMax=self.mx, minXRange=10000, maxXRange=10001, yMin=self.a_v[0], yMax=self.a_v[1], minYRange=self.a_v[1] - self.a_v[0], maxYRange=self.a_v[1] - self.a_v[0] + 1)
        if not self.cl_or_wh_c:
            if self.cur_g_y:
                self.a_c = self.cur_min_n_max(self.cur_g_y[-1])
            else:
                self.a_c = [0, 1]
            self.sc2.setLimits(xMin=self.mi, xMax=self.mx, minXRange=10000, maxXRange=10001, yMin=self.a_c[0], yMax=self.a_c[1], minYRange=self.a_c[1] - self.a_c[0], maxYRange=self.a_c[1] - self.a_c[0] + 1)
            
    #스크롤바
    
    def hsb_1_action(self):
        self.cl_or_wh_v = True
        self.autoRange_1.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.vr = self.sc1.ret_vr()
        #화면에 나온 분량 중간값
        self.hhalf1 = (self.vr[0][1] - self.vr[0][0]) / 2
        scroll_x = self.hsb_1.value()
        left = int(scroll_x - self.hhalf1)
        right = int(scroll_x + self.hhalf1)
        if left < 0:
            left = 0
            right = int(self.hhalf1*2)
        if left == 0:
            right = int(self.vr[0][1] - self.vr[0][0])
        elif right > self.mx:
            right = int(self.mx)
            left = int(self.mx - self.vr[0][1] - self.vr[0][0])
        
        self.sc1.setLimits(xMin=left, xMax=right, minXRange = self.vr[0][1] - self.vr[0][0], maxXRange = self.vr[0][1] - self.vr[0][0] + 0.001)
        self.draw_v_gr()
        
        
    def hsb_2_action(self):
        self.cl_or_wh_c = True
        self.autoRange_2.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.vr = self.sc2.ret_vr()
        #화면에 나온 분량 중간값
        self.hhalf2 = (self.vr[0][1] - self.vr[0][0]) / 2
        scroll_x = self.hsb_2.value()
        left = int(scroll_x - self.hhalf2)
        right = int(scroll_x + self.hhalf2)
        if left < 0:
            left = 0
            right = int(self.hhalf2*2)
        if left == 0:
            right = int(self.vr[0][1] - self.vr[0][0])
        elif right > self.mx:
            right = int(self.mx)
            left = int(self.mx - self.vr[0][1] - self.vr[0][0])
        
        self.sc2.setLimits(xMin=left, xMax=right, minXRange = self.vr[0][1] - self.vr[0][0], maxXRange = self.vr[0][1] - self.vr[0][0] + 0.001)
        self.draw_c_gr()
        
    def vsb_1_action(self):
        self.cl_or_wh_v = True
        self.autoRange_1.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.vr = self.sc1.ret_vr()
        #화면에 나온 분량 중간값
        self.vhalf1 = (self.vr[1][1] - self.vr[1][0]) / 2
        scroll_y = self.vsb_1.value()
        down = int(scroll_y - self.vhalf1)
        up = int(scroll_y + self.vhalf1)
        if down <= 0:
            down = 0
            up = int(self.vr[1][1] - self.vr[1][0])
        if up >= 85:
            up = 85
            down = 85 - int(self.vr[1][1] - self.vr[1][0])
        
        self.sc1.setLimits(yMin=down, yMax=up, minYRange=self.vr[1][1] - self.vr[1][0], maxYRange=self.vr[1][1] - self.vr[1][0] + 0.001)
        self.draw_v_gr()
        
    def vsb_2_action(self):
        self.cl_or_wh_c = True
        self.autoRange_2.setStyleSheet("background-color: rgb(50, 50, 50);\ncolor: rgb(255, 255, 255);")
        self.vr = self.sc2.ret_vr()
        #화면에 나온 분량 중간값
        self.vhalf2 = (self.vr[1][1] - self.vr[1][0]) / 2
        scroll_y = self.vsb_2.value()
        down = int(scroll_y - self.vhalf2)
        up = int(scroll_y + self.vhalf2*2)
        if down <= 0:
            down = 0
            up = int(self.vr[1][1] - self.vr[1][0])
        if up >= 165:
            up = 165
            down = 165 - int(self.vr[1][1] - self.vr[1][0])
        
        self.sc2.setLimits(yMin=down, yMax=up, minYRange=self.vr[1][1] - self.vr[1][0], maxYRange=self.vr[1][1] - self.vr[1][0] + 0.001)
        self.draw_c_gr()
        
    
    #연결
    def cnt(self, port_name):
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        if port_name:
            self.c = int(port_name[-1])
            self.serial.open(QIODevice.ReadWrite)
            #현재 IRANGE 확인
            self.serial.write("?IRANGE:\r".encode())
            
            self.s = self.serial.readLine()
            self.a= ''
            for i in self.s:
                self.a = self.a + i.decode()
            if self.a != '':
                self.range = int(self.a[8])
            
            #현재 ARPT 확인
            self.serial.write("?ARPT:\r".encode())
            
            return True
        else:
            return False
    
    #arpt 확인
    def ck_arpt(self):
        self.s = self.serial.readLine()
        self.a= ''
        self.b = ''
        for i in self.s:
            self.a = self.a + i.decode()
        if self.a != '':
            i = 6
            for d in self.a:
                if i > 0:
                    i -= 1
                    continue
                if ord(d) >= 48 and ord(d) <= 57:
                    self.b += d
                else:
                    break
        if self.b:
            self.arpt = int(self.b)
        
    #연결 확인
    def srl_IsOpen(self, port_name):
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        return self.serial.isOpen()
    
    #연결 해제
    def disconnect_serial(self):
        #정보 받아오는 쓰레드 중지
        self.re.stop()
        self.ts.stop()
        #연결 해제
        self.c = -1
        self.tm = 0
        print("연결 해제됨")
        return self.serial.close()
    
    #커멘드 입력
    def rPressed(self):
        self.s = self.command.text()
        self.st = self.s + '\r'
        self.serial.write(self.st.encode())
        self.s = (str)("<<<" + self.s)
        self.board_c.append(self.s)
        self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
        self.command.clear()
    
    #버튼 클릭
    def btn_clicked(self):
        self.btn_value = self.sender().text()
        #Connect
        if self.btn_value == 'Connect':
            #기존 연결 해제
            self.disconnect_serial()
            #연결하기
            self.con = self.cnt(self.cbPort.currentText())
            if self.con:
                #정보 받아오는 쓰레드 생성
                self.re = Receive(self)
                self.re.start()
            print("연결 성공")
        
        #Disconnect
        elif self.btn_value == 'Disconnect':
            self.disconnect_serial()
            self.tm = 0
        
        #Quit
        elif self.btn_value == 'Quit':
            self.lb.raise_()
            self.lb.setHidden(False)
            self.lb.repaint()
            self.disconnect_serial()
            #ck가 멈출 때까지 대기
            #포트 체크 중지
            self.ck.stop()
            QCoreApplication.instance().quit()
        
        #?REV:, ?SN:, ?IRANGE:, ?MEAS:
        elif self.btn_value == '?REV:' or self.btn_value == '?SN:' or self.btn_value == '?IRANGE:' or self.btn_value == '?MEAS:':
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        
        #SET IRANGE 0
        elif self.btn_value == 'IRANGE:0':
            self.range = 0
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #SET IRANGE 1
        elif self.btn_value == 'IRANGE:1':
            self.range = 1
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
        
        #ARPT:0
        elif self.btn_value == 'ARPT:0':
            self.arpt = 0
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #ARPT:50
        elif self.btn_value == 'ARPT:50':
            self.arpt = 50
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #ARPT:1000
        elif self.btn_value == 'ARPT:1000':
            self.arpt = 1000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #ARPT:5000
        elif self.btn_value == 'ARPT:5000':
            self.arpt = 5000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #ARPT:10000
        elif self.btn_value == 'ARPT:10000':
            self.arpt = 10000
            self.s = self.btn_value + '\r'
            self.serial.write(self.s.encode())
            self.s = (str)("<<<" + self.btn_value)
            self.board_c.append(self.s)
            self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
            self.command.clear()
            
        #AutoRange
        elif self.btn_value == 'V_AutoRange':
            self.autoRange_1.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
            self.cl_or_wh_v = False
        elif self.btn_value == 'C_AutoRange':
                self.autoRange_2.setStyleSheet("background-color: rgb(95, 255, 71);\ncolor: rgb(0, 0, 0);")
                self.cl_or_wh_c = False
            
    
    def saveExcel(self):
        ex = Excel(self)
        
        ex.create()
        ex.write()
        ex.save()
    
    def saveExcelGraph(self):
        ex = Excel(self)
        
        ex.create()
        ex.write()
        ex.save_graph()
        
    #멤버 변수 접근
    def add_tm(self, tm):
        self.tm += tm
    
    def get_tm(self):
        return self.tm
    
    def get_c(self):
        return self.c
    
    def get_currentText(self):
        return self.cbPort.currentText()
    
    def cb_remove(self, item):
        self.cbPort.removeItem(item)
        
    def cb_findText(self, port_name):
        return self.cbPort.findText(port_name)
    
    def cb_count(self):
        return self.cbPort.count()
    
    def cb_addItem(self, item):
        self.cbPort.addItem(item)
        
    def cb_itemText(self, text):
        return self.cbPort.itemText(text)
    
    def cb_insertItem(self, num, name):
        self.cbPort.insertItem(num, name)
    
    def get_serial_readLine(self):
        return self.serial.readLine()
    
    def get_board_c(self):
        return self.board_c
    
    def get_btn_value(self):
        return self.btn_value
    
    def get_VOL(self):
        return self.VOL
    
    def get_CUR(self):
        return self.CUR
    
    def get_arpt(self):
        return self.arpt
    
    def get_g(self):
        return self.g
    
    def get_vol_g_y(self):
        return self.vol_g_y
    
    def get_cur_g_y(self):
        return self.cur_g_y
    
    def g_append(self):
        self.g.append(self.tm)
        
    def vol_g_y_append(self, vol):
        self.vol_g_y.append(vol)
        
    def cur_g_y_append(self, cur):
        self.cur_g_y.append(cur)
        
    def board_c_append(self, a):
        self.board_c.append(a)
        
    def set_verScroll(self):
        self.board_c.verticalScrollBar().setValue(self.board_c.verticalScrollBar().maximum())
    
    def VOL_setText(self, text):
        self.VOL.setText(text)
        
    def CUR_setText(self, text):
        self.VOL.setText(text)

#들어오는 정보 감지 Thread
class Receive(QThread):
    def __init__(self, parent):
        QThread.__init__(self)
        self._status = False
        self.serial = serial
        self.re_w = True
        super().__init__(parent)
        self.parent = parent
        self.serial = QSerialPort()
        info = QSerialPortInfo(self.parent.get_currentText())
        self.serial.setPort(info)
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        
    #그래프 요소 추가
    def setGraph(self, vol, cur):
        self.parent.g_append()
        self.parent.vol_g_y_append(vol)
        self.parent.cur_g_y_append(cur)
        
    def run(self):
        while self.re_w:
            self.s = self.parent.get_serial_readLine()
            if self.s:
                self.a= '>'
                for i in self.s:
                    self.a = self.a + i.decode('ascii')
                self.parent.board_c_append(self.a)
                self.parent.set_verScroll()
                if self.parent.get_btn_value() == '?MEAS:':
                    self.temp = self.a[5:]
                    self.tmp = self.temp.split(",")
                    
                    #vol display
                    while len(self.tmp[2]) < 6:
                        self.tmp[2] = self.tmp[2] + '0'
                    
                    self.parent.VOL_setText(self.tmp[2][:3] + "," + self.tmp[2][3:])
                    
                    #cur display
                    while len(self.tmp[1]) < 6:
                        self.tmp[1] = self.tmp[1] + '0'
                    
                    self.parent.CUR_setText(self.tmp[1][:3] + "," + self.tmp[1][3:])
                    
                    
                if self.a[:6] == '>>ARPT' and self.parent.get_arpt() != 0:
                    self.temp = self.a[7:]
                    self.tmp = self.temp.split(",")

                    #그래프 X축 범위 늘림
                    self.parent.add_tm(self.parent.get_arpt())
                    #그래프 추가
                    self.setGraph(float(self.tmp[3]) / 1000, float(self.tmp[2]) / 1000)
                    #그래프 그리기
                    self.parent.draw_v_gr()
                    self.parent.draw_c_gr()
                    
            QTest.qWait(50)
            
    def stop(self):
        '''
        QThread 종료
        '''
        self.re_w = False
        
        self.wait()
        self.quit()

#포트 감지 Thread
class Check(QThread):
    def __init__(self, parent):
        QThread.__init__(self, parent)
        self._status = False
        self.serial = serial
        self.ck_w = True
        super().__init__(parent)
        self.parent = parent
    
    @staticmethod
    def get_port_path():
        """
        현재플래폼에 맞게 경로 또는 지정어를 반환
        :return:
        """
        return {"linux": '/dev/ttyS', "win32": 'COM'}[__platform__]
    
    # baudrate= 115200, data_bits=8, flow_control=None, parity=None, stop_bits=1
    def _open(self, port_name):
        """
        포트 연결 설정
        """
        self.serial = QSerialPort()
        info = QSerialPortInfo(port_name)
        self.serial.setPort(info)
        self.serial.setBaudRate(QSerialPort.Baud115200)
        self.serial.setDataBits(QSerialPort.Data8)
        self.serial.setFlowControl(QSerialPort.NoFlowControl)
        self.serial.setParity(QSerialPort.NoParity)
        self.serial.setStopBits(QSerialPort.OneStop)
        return self.serial.open(QIODevice.ReadWrite)
    
    def run(self):
        """
        255개의 포트를 열고 닫으면서 사용가능한 포트를 찾아서 입력
        """
        while self.ck_w:
            port_path = 'COM'
            for number in range(255):
                #ex) COM3
                port_name = port_path + str(number)
                #열리지 않고 이미 열려있지 않은 포트 삭제
                if (not self._open(port_name)) and number != self.parent.get_c():
                    self.parent.cb_remove(self.parent.cb_findText(port_name))
                    continue
                #포트가 열린다면 정렬 후 삽입
                if self._open(port_name):
                    #정렬
                    if self.parent.cb_count() != 0:
                        for num in range(0, self.parent.cb_count()):
                            self.tmp = self.parent.cb_itemText(num)
                            if int(self.tmp[-1]) > num:
                                self.parent.cb_insertItem(num+1, port_name)
                                
                    elif self.parent.cb_count() == 0:
                        self.parent.cb_addItem(port_name)
                    self.serial.close()
            #sleep
            time.sleep(1)
        
    def stop(self):
        '''
        QThread 종료
        '''
        self.ck_w = False
        self.wait()
        self.quit()
    
#작동 확인용 쓰레드
class Test(QThread):
    def __init__(self, parent):
        QThread.__init__(self)
        self._status = False
        self.serial = serial
        self.ck_w = True
        super().__init__(parent)
        self.parent = parent
        self.se = True
        
    def setGraph(self, vol, cur):
        self.parent.get_g().append(self.parent.get_tm())
        self.parent.get_vol_g_y().append(vol)
        self.parent.get_cur_g_y().append(cur)
    
    def run(self):
        while self.se:
            self.parent.tm += 50
            self.setGraph(random.randrange(0, 86), random.randrange(0, 166))
            #그래프 그리기
            self.parent.draw_v_gr()
            self.parent.draw_c_gr()
            
            QTest.qWait(50)
            
    def stop(self):
        '''
        QThread 종료
        '''
        self.se = False
        self.wait()
        self.quit()
    

class Excel():
    def __init__(self, parent):
        self.parent = parent
        self.today = datetime.now()
        self.wb = None
    
    def create(self):
        #생성
        self.wb = openpyxl.Workbook()
        #setting sheet name
        self.wb.active.title = "ezDAQ-Monitor"
        #setting sheet location
        self.w1 = self.wb["ezDAQ-Monitor"]
        #setting column name
        self.w1.cell(1, 1).value = 'ms'
        self.w1.cell(1, 2).value = 'v'
        self.w1.cell(1, 3).value = 'mA'
        
    def write(self):
        for t in range(0, len(self.parent.g)):
            self.w1.cell(t+2, 1).value = self.parent.g[t]
            self.w1.cell(t+2, 2).value = self.parent.vol_g_y[t]
            self.w1.cell(t+2, 3).value = self.parent.cur_g_y[t]
        
    #저장
    def save(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.new__filename = "C:\ezDAQ-Monitor\excel\Excel" + today_s + ".xlsx"
        
        self.wb.save(self.new__filename)
    
    #그래프와 함께 저장
    def save_graph(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.new__filename = "C:\ezDAQ-Monitor\excel\Excel_Graph" + today_s + ".xlsx"
        self.rd = self.wb
        sheet_list = self.rd.sheetnames
        s0 = self.rd[sheet_list[0]]
        
        chart = LineChart()
        #Data 범위 설정
        xval = Reference(s0, min_col=1, max_col=1, min_row=2, max_row=len(self.parent.g)+2)
        yval = Reference(s0, min_col = 2, max_col =2, min_row = 2, max_row = len(self.parent.g)+2)
        chart.add_data(yval, titles_from_data=True)
        chart.set_categories(xval)
        chart.style = 1
        
        chart.height = 10
        chart.width = (len(self.parent.g)+2)
        s0.add_chart(chart, "E5")
        
        self.rd.save(self.new__filename)
        
    '''
    def read(self):
        today_s = self.today.strftime('%m_%d__%H_%M_%S')
        self.filename = "C:\ezDAQ-Monitor\excel\Excel" + today_s + ".xlsx"
        self.rd = openpyxl.load_workbook(self.filename, data_only=True)
        sheet_list = self.rd.sheetnames
        s0 = self.rd[sheet_list[0]]
        
        chart = LineChart()
        #Data 범위 설정
        xval = Reference(s0, min_col=1, max_col=1, min_row=2, max_row=len(self.parent.g)+2)
        yval = Reference(s0, min_col = 2, max_col =2, min_row = 2, max_row = len(self.parent.g)+2)
        chart.add_data(yval, titles_from_data=True)
        chart.set_categories(xval)
        chart.style = 1
        
        chart.height = 10
        chart.width = (len(self.parent.g)+2)
        s0.add_chart(chart, "E5")
        
        self.new_filename = "C:\ezDAQ-Monitor\excel\Excel_Graph" + today_s + ".xlsx"
        self.rd.save(self.new_filename)
'''
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    myApp = MyApp()
    myApp.show()
    
    app.exec_()